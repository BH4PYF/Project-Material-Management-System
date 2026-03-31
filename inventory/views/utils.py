import json
import logging
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from functools import wraps

from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import JsonResponse
from django.db.models import Max
from django.db.models.functions import Substr, Cast
from django.db.models import IntegerField
from django.utils import timezone

from ..models import (
    Profile, InboundRecord, OperationLog,
)

logger = logging.getLogger('inventory')

# ========== 工具函数 ==========


def is_ajax_request(request):
    """统一判断 AJAX/JSON 请求。"""
    return request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.content_type == 'application/json'


def log_operation(user, module, op_type, details, related_no=''):
    OperationLog.objects.create(
        operator=user.username if hasattr(user, 'username') else str(user),
        module=module, op_type=op_type, details=details, related_no=related_no,
    )


def is_admin(user):
    return hasattr(user, 'profile') and user.profile.is_admin


def is_material_dept(user):
    return hasattr(user, 'profile') and user.profile.is_material_dept


def is_clerk(user):
    return hasattr(user, 'profile') and user.profile.is_clerk


def is_supplier(user):
    """判断是否为供应商"""
    return hasattr(user, 'profile') and user.profile.is_supplier


def can_manage_inventory(user):
    """判断用户是否可以管理入库记录（管理员、物资部、材料员）"""
    if not hasattr(user, 'profile'):
        return False
    return user.profile.is_admin or user.profile.is_material_dept or user.profile.is_clerk


def can_manage_purchase_plan(user):
    """判断用户是否可以管理采购计划（管理员、物资部、材料员）"""
    if not hasattr(user, 'profile'):
        return False
    return user.profile.is_admin or user.profile.is_material_dept or user.profile.is_clerk


def can_manage_delivery(user):
    """判断用户是否可以管理发货（管理员、物资部、供应商）"""
    if not hasattr(user, 'profile'):
        return False
    return user.profile.is_admin or user.profile.is_material_dept or user.profile.is_supplier


def _permission_required(check_fn, deny_message=None):
    """通用权限装饰器工厂。

    check_fn:     接收 user 返回 bool 的权限判定函数。
    deny_message: 可选拒绝提示。提供时会通过 messages.error 显示给用户；
                  未提供时仅静默重定向（保持 admin_required 原有行为）。
    """
    def decorator(view_func):
        @wraps(view_func)
        @login_required
        def wrapper(request, *args, **kwargs):
            if not check_fn(request.user):
                if is_ajax_request(request):
                    return JsonResponse({'error': deny_message or '无权限'}, status=403)
                if deny_message:
                    from django.contrib import messages
                    messages.error(request, deny_message)
                return redirect('dashboard')
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


admin_required = _permission_required(is_admin)
inventory_required = _permission_required(can_manage_inventory, '无权限访问入库管理')
purchase_plan_required = _permission_required(can_manage_purchase_plan, '无权限访问采购计划')
delivery_required = _permission_required(can_manage_delivery, '无权限访问发货管理')
supplier_required = _permission_required(is_supplier, '仅限供应商操作')


def role_required(*roles):
    """通用角色权限装饰器，支持传入多个角色名"""
    def decorator(view_func):
        @wraps(view_func)
        @login_required
        def wrapper(request, *args, **kwargs):
            if not hasattr(request.user, 'profile') or request.user.profile.role not in roles:
                if is_ajax_request(request):
                    return JsonResponse({'error': '无权限'}, status=403)
                return redirect('dashboard')
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


def permission_required(perm):
    """Django权限装饰器，验证Django权限系统的权限"""
    def decorator(view_func):
        @wraps(view_func)
        @login_required
        def wrapper(request, *args, **kwargs):
            if not request.user.has_perm(perm):
                if is_ajax_request(request):
                    return JsonResponse({'error': '无权限执行此操作'}, status=403)
                return redirect('dashboard')
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


def combined_permission_required(perm=None, roles=None):
    """组合权限装饰器，同时验证Django权限和角色权限
    
    Args:
        perm: Django权限字符串，如 'inventory.add_material'
        roles: 角色列表，如 ['admin', 'material_dept']
    """
    def decorator(view_func):
        @wraps(view_func)
        @login_required
        def wrapper(request, *args, **kwargs):
            # 验证Django权限（如果提供）
            if perm and not request.user.has_perm(perm):
                if is_ajax_request(request):
                    return JsonResponse({'error': '无权限执行此操作'}, status=403)
                return redirect('dashboard')
            
            # 验证角色权限（如果提供）
            if roles:
                if not hasattr(request.user, 'profile') or request.user.profile.role not in roles:
                    if is_ajax_request(request):
                        return JsonResponse({'error': '无权限'}, status=403)
                    return redirect('dashboard')
            
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


def generate_code(prefix, model_class, field='code'):
    """生成唯一编号，带重试机制以应对并发冲突。

    在高并发场景下 select_for_update 仅锁定已有行，无法阻止其他事务
    同时插入新行。因此此函数在 IntegrityError 时自动重试。
    """
    from django.db import transaction, IntegrityError

    # 使用 all_objects（含软删除记录）避免编码冲突，
    # 若模型无 all_objects 则回退到默认 objects
    manager = getattr(model_class, 'all_objects', model_class.objects)

    for attempt in range(3):
        with transaction.atomic():
            qs = manager.select_for_update().filter(
                **{f'{field}__regex': rf'^{prefix}\d{{4}}$'}
            )
            last = qs.order_by(f'-{field}').first()
            if last:
                val = getattr(last, field)
                suffix = val[len(prefix):]
                next_num = int(suffix) + 1 if suffix.isdigit() else 1
            else:
                next_num = 1
            code = f"{prefix}{next_num + attempt:04d}"
            # 验证生成的编号是否已存在（含软删除记录）
            if not manager.filter(**{field: code}).exists():
                return code
    # 兜底：使用时间戳保证唯一性
    return f"{prefix}{timezone.now().strftime('%Y%m%d%H%M%S')}"


def save_with_generated_code(obj, prefix, model_class):
    """生成编号并保存对象，自动处理唯一约束冲突重试。

    成功时返回 True，3 次重试均失败时返回 False。
    """
    from django.db import transaction, IntegrityError

    last_generated_code = None
    for attempt in range(3):
        try:
            with transaction.atomic():
                if last_generated_code:
                    num_part = int(last_generated_code[len(prefix):])
                    obj.code = f"{prefix}{num_part + 1:04d}"
                else:
                    obj.code = generate_code(prefix, model_class)
                obj.save()
                return True
        except IntegrityError as e:
            if 'UNIQUE constraint failed' in str(e):
                last_generated_code = obj.code
            else:
                raise
    return False


def generate_no(prefix, model_class=None):
    """生成唯一编号（使用 select_for_update 保证并发安全）
    prefix: 前缀 (如 'IN', 'PP')
    model_class: 查询的模型类，默认为 InboundRecord

    注意：调用方必须在 transaction.atomic() 块内调用此函数，
    以确保 select_for_update 的锁覆盖到后续的 .save() 操作。
    """
    if model_class is None:
        model_class = InboundRecord

    today = timezone.now().strftime('%Y%m%d')
    full_prefix = f"{prefix}{today}"
    prefix_len = len(full_prefix)

    # 查找今天最大的序号（包含软删除记录，避免编号冲突）
    # select_for_update 锁由调用方的 atomic 块持有
    queryset = model_class.all_objects if hasattr(model_class, 'all_objects') else model_class.objects
    # 使用数据库 MAX + Substr/Cast 聚合，避免将所有记录加载到 Python
    max_num = queryset.select_for_update().filter(
        no__startswith=full_prefix
    ).annotate(
        num_part=Cast(Substr('no', prefix_len + 1), IntegerField())
    ).aggregate(max_num=Max('num_part'))['max_num'] or 0

    next_num = max_num + 1
    return f"{full_prefix}{next_num:04d}"


def get_supplier_display_name(supplier):
    """安全获取供应商显示名称，避免链式属性访问崩溃
    
    参数可以是：
    - Supplier 对象：直接返回供应商名称
    - User 对象：通过 profile.supplier_info 获取供应商
    """
    from ..models import Supplier
    
    # 如果是 Supplier 对象，直接返回名称
    if isinstance(supplier, Supplier):
        return supplier.name or supplier.code or '未知供应商'
    
    # 如果是 User 对象，尝试获取关联的供应商
    try:
        profile = supplier.profile
        if profile and profile.supplier_info:
            return profile.supplier_info.name or profile.supplier_info.code or '未知供应商'
        # 没有关联供应商，返回用户信息
        return supplier.first_name or supplier.username
    except (AttributeError, Profile.DoesNotExist):
        return getattr(supplier, 'first_name', None) or getattr(supplier, 'username', '未知供应商')


def get_supplier_from_user(user):
    """安全获取用户关联的供应商档案对象，无关联时返回 None"""
    try:
        profile = user.profile
        if profile.supplier_info_id:
            return profile.supplier_info
    except (AttributeError, Profile.DoesNotExist):
        pass
    return None


def make_attachment_disposition(filename):
    """生成兼容中文文件名的 Content-Disposition 头值。
    同时提供 filename（ASCII 回退）和 filename*（RFC 5987 UTF-8 编码）。
    """
    from urllib.parse import quote
    ascii_name = filename.encode('ascii', 'replace').decode()
    encoded = quote(filename)
    return f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{encoded}"


def decimal_default(obj):
    if isinstance(obj, Decimal):
        return float(obj)
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    raise TypeError


def validate_required_fields(post_data, field_map):
    """检查必填字段，返回第一个缺失字段的错误信息，全部通过时返回 None。

    post_data: request.POST (QueryDict)
    field_map: {field_name: error_message, ...}
               按插入顺序检查，遇到第一个空值即返回对应 error_message。
    """
    for field, error_msg in field_map.items():
        val = post_data.get(field, '')
        if isinstance(val, str):
            val = val.strip()
        if not val:
            return error_msg
    return None


def create_user_for_supplier(supplier, password='12345678'):
    """为供应商自动创建关联的用户账号。

    用户名使用供应商编号（小写），如 sup0001。
    如果该供应商已有关联用户则跳过，返回 (None, '已存在')。
    如果用户名已被占用则跳过，返回 (None, '用户名冲突')。
    成功时返回 (user, None)。
    """
    # 已有关联用户，跳过
    if supplier.user_profiles.exists():
        return None, '已存在'

    username = supplier.code.lower()

    # 用户名冲突
    if User.objects.filter(username=username).exists():
        return None, '用户名冲突'

    user = User.objects.create_user(
        username=username,
        password=password,
        first_name=supplier.name[:30],
    )
    Profile.objects.update_or_create(
        user=user,
        defaults={
            'role': 'supplier',
            'phone': supplier.phone,
            'supplier_info': supplier,
        },
    )
    return user, None


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    val = str(val).strip()
    if not val:
        return None
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d'):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    val = str(val).strip()
    if not val:
        return None
    try:
        return Decimal(val)
    except InvalidOperation:
        return None


def parse_positive_decimal(val, field_name, allow_zero=False):
    """解析并验证正数 Decimal，返回 (value, error_msg)。
    成功时 error_msg 为 None；失败时 value 为 None。
    """
    d = _parse_decimal(val)
    if d is None:
        return None, f'{field_name}格式不正确'
    if allow_zero:
        if d < 0:
            return None, f'{field_name}不能为负数'
    else:
        if d <= 0:
            return None, f'{field_name}必须为正数'
    return d, None


# ========== Excel 导出工具 ==========


def create_excel_workbook(title, headers, style='primary'):
    """创建带格式化表头的 Excel 工作簿。

    style:
      'primary' — 蓝底白字（发货单、采购计划、入库记录等列表导出）
      'report'  — 灰底黑字 + 边框（报表类导出）

    返回 (wb, ws, border)。border 仅在 style='report' 时非 None。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = title

    if style == 'primary':
        fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        font = Font(bold=True, color='FFFFFF', size=12)
        alignment = Alignment(horizontal='center', vertical='center')
        border = None
    else:
        fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        font = Font(bold=True)
        alignment = None
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = font
        cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border

    return wb, ws, border


def set_column_widths(ws, widths):
    """批量设置列宽。widths 为宽度列表，依次对应第 1、2、... 列。"""
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def make_excel_response(wb, filename):
    """将 openpyxl Workbook 写入 HttpResponse 并设置下载头。"""
    from django.http import HttpResponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = make_attachment_disposition(filename)
    wb.save(response)
    return response
