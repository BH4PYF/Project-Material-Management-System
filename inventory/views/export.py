from decimal import Decimal, InvalidOperation
import logging

from django.db import transaction, IntegrityError, DatabaseError
from django.db.models import Sum
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_POST

from ..models import Material, InboundRecord, Category, Supplier
from .utils import (
    admin_required, log_operation, parse_date, role_required,
    generate_code,
    create_excel_workbook, set_column_widths, make_excel_response,
)

logger = logging.getLogger('inventory')


@role_required('admin', 'material_dept', 'clerk')
def export_excel(request):
    export_type = request.GET.get('type', 'inventory')
    
    # 导出数量限制，防止数十万条数据导致 OOM
    MAX_EXPORT_ROWS = 10000  # 最大导出 1 万条

    if export_type == 'inventory':
        headers = ['材料编号', '材料名称', '分类', '规格', '单位', '累计入库量', '安全库存', '入库均价', '入库总值']
        wb, ws, _ = create_excel_workbook('入库汇总', headers, style='report')

        # 限制查询数量
        materials_qs = Material.objects.select_related('category').all()[:MAX_EXPORT_ROWS]

        # 批量聚合入库数据，一次查询替代 N*3 次查询
        inbound_agg = InboundRecord.objects.filter(
            material_id__in=materials_qs.values_list('pk', flat=True)
        ).values('material_id').annotate(
            total_qty=Sum('quantity'),
            total_amount=Sum('total_amount'),
        )
        agg_map = {
            row['material_id']: {
                'total_qty': row['total_qty'] or Decimal('0'),
                'total_amount': row['total_amount'] or Decimal('0'),
            }
            for row in inbound_agg
        }

        row = 2
        for m in materials_qs.iterator(chunk_size=500):
            agg = agg_map.get(m.pk, {'total_qty': Decimal('0'), 'total_amount': Decimal('0')})
            inbound_qty = agg['total_qty']
            inbound_amount = agg['total_amount']
            avg_cost = (inbound_amount / inbound_qty) if inbound_qty > 0 else Decimal('0')
            inbound_value = inbound_amount if inbound_qty > 0 else Decimal('0')

            ws.cell(row=row, column=1, value=m.code)
            ws.cell(row=row, column=2, value=m.name)
            ws.cell(row=row, column=3, value=m.category.name)
            ws.cell(row=row, column=4, value=m.spec)
            ws.cell(row=row, column=5, value=m.unit)
            ws.cell(row=row, column=6, value=float(inbound_qty))
            ws.cell(row=row, column=7, value=float(m.safety_stock))
            ws.cell(row=row, column=8, value=float(avg_cost))
            ws.cell(row=row, column=9, value=float(inbound_value))
            row += 1

        set_column_widths(ws, [12, 20, 12, 15, 8, 12, 10, 10, 12])

        filename = '入库汇总.xlsx'

    elif export_type == 'inbound':
        headers = ['入库单号', '日期', '项目', '项目地址', '材料', '单位', '规格', '数量', '单价', '总金额', '供应商']
        wb, ws, _ = create_excel_workbook('入库记录', headers, style='primary')

        # 使用 select_related 优化关联查询（SoftDeleteManager 已自动过滤已删除记录）
        records = InboundRecord.objects.select_related('project', 'material', 'supplier')

        date_from = parse_date(request.GET.get('date_from'))
        date_to = parse_date(request.GET.get('date_to'))
        if date_from:
            records = records.filter(date__gte=date_from)
        if date_to:
            records = records.filter(date__lte=date_to)

        project_id = request.GET.get('project')
        if project_id:
            records = records.filter(project_id=project_id)

        material_id = request.GET.get('material')
        if material_id:
            records = records.filter(material_id=material_id)

        supplier_id = request.GET.get('supplier')
        if supplier_id:
            records = records.filter(supplier_id=supplier_id)

        # 过滤完成后再限制数量，使用 iterator 分块加载避免全量缓存
        records = records.order_by('date', 'no')[:MAX_EXPORT_ROWS]

        row = 2
        for r in records.iterator(chunk_size=500):
            ws.cell(row=row, column=1, value=r.no)
            ws.cell(row=row, column=2, value=str(r.date))
            ws.cell(row=row, column=3, value=f"{r.project.code} - {r.project.name}")
            ws.cell(row=row, column=4, value='-')
            ws.cell(row=row, column=5, value=r.material.name)
            ws.cell(row=row, column=6, value=r.material.unit)
            ws.cell(row=row, column=7, value=r.spec or '-')
            ws.cell(row=row, column=8, value=float(r.quantity))
            ws.cell(row=row, column=9, value=float(r.unit_price))
            ws.cell(row=row, column=10, value=float(r.total_amount))
            ws.cell(row=row, column=11, value=r.supplier.name)
            row += 1

        set_column_widths(ws, [15, 12, 20, 25, 15, 8, 15, 10, 10, 12, 20])

        filename = f'入库记录_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    else:
        return JsonResponse({'error': '无效的导出类型'}, status=400)

    log_operation(request.user, 'Excel 导出', 'export', f'导出{export_type}数据（最多{MAX_EXPORT_ROWS}条）')
    return make_excel_response(wb, filename)


# ========== 导入模板下载 ==========

MATERIAL_IMPORT_HEADERS = ['材料名称', '分类名称', '规格型号', '计量单位', '标准单价', '备注']
SUPPLIER_IMPORT_HEADERS = ['供应商名称', '联系人', '联系电话', '地址', '主营材料类型', '信用等级(优秀/良好/一般)', '合作开始日期(YYYY-MM-DD)', '备注']

MATERIAL_UNITS = [c[0] for c in Material.UNIT_CHOICES]
SUPPLIER_CREDIT_MAP = {'优秀': 'excellent', '良好': 'good', '一般': 'average'}


@admin_required
def download_import_template(request):
    """下载导入模板"""
    tpl_type = request.GET.get('type', 'material')

    if tpl_type == 'material':
        wb, ws, _ = create_excel_workbook('材料导入模板', MATERIAL_IMPORT_HEADERS, style='primary')
        sample = ['螺纹钢 HRB400', '钢材', 'Φ12mm', '吨', 4200, '']
        for col, val in enumerate(sample, 1):
            ws.cell(row=2, column=col, value=val)
        set_column_widths(ws, [20, 12, 15, 10, 12, 20])
        filename = '材料导入模板.xlsx'
    elif tpl_type == 'supplier':
        wb, ws, _ = create_excel_workbook('供应商导入模板', SUPPLIER_IMPORT_HEADERS, style='primary')
        sample = ['华东钢铁有限公司', '陈经理', '13800138001', '上海市宝山区', '钢材', '优秀', '2024-01-01', '']
        for col, val in enumerate(sample, 1):
            ws.cell(row=2, column=col, value=val)
        set_column_widths(ws, [25, 12, 15, 25, 15, 18, 20, 20])
        filename = '供应商导入模板.xlsx'
    else:
        return JsonResponse({'error': '无效的模板类型'}, status=400)

    return make_excel_response(wb, filename)


# ========== Excel 导入 ==========

def _read_excel_rows(file, max_size=10 * 1024 * 1024):
    """读取上传的 Excel 文件，返回 (rows, error_msg)"""
    if file.size > max_size:
        return None, '文件过大，最大支持 10MB'
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
        return rows, None
    except (KeyError, IndexError, TypeError, ValueError) as e:
        logger.exception('无法解析 Excel 文件: %s', e)
        return None, '无法解析文件，请确保为 .xlsx 格式的 Excel 文件'


@admin_required
@require_POST
def import_excel(request):
    """从 Excel 批量导入材料或供应商"""
    import_type = request.POST.get('type', '')
    file = request.FILES.get('file')
    if not file:
        return JsonResponse({'error': '请选择文件'}, status=400)

    rows, err = _read_excel_rows(file)
    if err:
        return JsonResponse({'error': err}, status=400)

    if not rows:
        return JsonResponse({'error': 'Excel 文件中没有数据行'}, status=400)

    if import_type == 'material':
        return _import_materials(request, rows)
    elif import_type == 'supplier':
        return _import_suppliers(request, rows)
    else:
        return JsonResponse({'error': '无效的导入类型'}, status=400)


def _import_materials(request, rows):
    """导入材料数据"""
    created = 0
    skipped = 0
    errors = []
    MAX_ROWS = 5000

    cat_map = {c.name: c for c in Category.objects.all()}

    try:
        with transaction.atomic():
            for i, row in enumerate(rows[:MAX_ROWS], start=2):
                if not row or not row[0]:
                    continue
                name = str(row[0]).strip()
                cat_name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                spec = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                unit = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                price_raw = row[4] if len(row) > 4 and row[4] else 0
                remark = str(row[5]).strip() if len(row) > 5 and row[5] else ''

                if not name:
                    errors.append(f'第 {i} 行：材料名称为空，已跳过')
                    skipped += 1
                    continue
                if not cat_name or cat_name not in cat_map:
                    errors.append(f'第 {i} 行：分类「{cat_name}」不存在，已跳过')
                    skipped += 1
                    continue
                if unit and unit not in MATERIAL_UNITS:
                    errors.append(f'第 {i} 行：单位「{unit}」不在系统支持列表中，已跳过')
                    skipped += 1
                    continue
                if not unit:
                    errors.append(f'第 {i} 行：计量单位为空，已跳过')
                    skipped += 1
                    continue

                try:
                    price = Decimal(str(price_raw))
                except (InvalidOperation, ValueError):
                    errors.append(f'第 {i} 行：单价或库存数值格式错误，已跳过')
                    skipped += 1
                    continue

                if Material.objects.filter(name=name, spec=spec).exists():
                    skipped += 1
                    continue

                code = generate_code('MAT', Material)
                Material.objects.create(
                    code=code, name=name, category=cat_map[cat_name],
                    spec=spec, unit=unit, standard_price=price,
                    safety_stock=0, remark=remark,
                )
                created += 1

    except (IntegrityError, DatabaseError, ValueError, KeyError) as e:
        logger.exception('材料导入失败: %s', e)
        return JsonResponse({'error': '导入过程中发生错误，所有数据已回滚'}, status=500)

    log_operation(request.user, '材料档案', 'create', f'Excel 批量导入材料：成功 {created} 条，跳过 {skipped} 条')
    return JsonResponse({
        'success': True,
        'message': f'导入完成：成功 {created} 条，跳过 {skipped} 条',
        'created': created,
        'skipped': skipped,
        'errors': errors[:20],
    })


def _import_suppliers(request, rows):
    """导入供应商数据"""
    created = 0
    skipped = 0
    errors = []
    MAX_ROWS = 5000

    cat_map = {c.name: c for c in Category.objects.all()}

    try:
        with transaction.atomic():
            for i, row in enumerate(rows[:MAX_ROWS], start=2):
                if not row or not row[0]:
                    continue
                name = str(row[0]).strip()
                contact = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                phone = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                address = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                main_type_name = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                credit_label = str(row[5]).strip() if len(row) > 5 and row[5] else '良好'
                start_date_raw = str(row[6]).strip() if len(row) > 6 and row[6] else ''
                remark = str(row[7]).strip() if len(row) > 7 and row[7] else ''

                if not name:
                    errors.append(f'第 {i} 行：供应商名称为空，已跳过')
                    skipped += 1
                    continue

                if Supplier.objects.filter(name=name).exists():
                    skipped += 1
                    continue

                main_type = cat_map.get(main_type_name)
                credit_rating = SUPPLIER_CREDIT_MAP.get(credit_label, 'good')

                start_date = None
                if start_date_raw:
                    start_date = parse_date(start_date_raw)

                code = generate_code('SUP', Supplier)
                Supplier.objects.create(
                    code=code, name=name, contact=contact, phone=phone,
                    address=address, main_type=main_type,
                    credit_rating=credit_rating, start_date=start_date,
                    remark=remark,
                )
                created += 1

    except (IntegrityError, DatabaseError, ValueError, KeyError) as e:
        logger.exception('供应商导入失败: %s', e)
        return JsonResponse({'error': '导入过程中发生错误，所有数据已回滚'}, status=500)

    log_operation(request.user, '供应商档案', 'create', f'Excel 批量导入供应商：成功 {created} 条，跳过 {skipped} 条')
    return JsonResponse({
        'success': True,
        'message': f'导入完成：成功 {created} 条，跳过 {skipped} 条',
        'created': created,
        'skipped': skipped,
        'errors': errors[:20],
    })
