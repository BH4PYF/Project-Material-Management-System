"""Celery 异步导出任务的单元测试"""
import os
from decimal import Decimal

import pytest
from django.test import TestCase, override_settings
from django.contrib.auth.models import User

from ..models import (
    Profile, Category, Material, Supplier,
    Project, InboundRecord, PurchasePlan, Delivery,
)
from ..tasks import (
    export_inventory_excel,
    export_inbound_excel,
    export_purchase_plans,
    export_deliveries,
)


TEMP_MEDIA = '/tmp/test_media_exports'


@override_settings(MEDIA_ROOT=TEMP_MEDIA)
class ExportTaskTestBase(TestCase):
    """导出任务测试的共享基础数据"""

    @classmethod
    def setUpTestData(cls):
        cls.user = User.objects.create_user(username='task_user', password='pass12345')
        Profile.objects.create(user=cls.user, role='admin')

        cls.category = Category.objects.create(name='测试分类', code='CAT_T')
        cls.material = Material.objects.create(
            name='测试材料', code='MAT_T', category=cls.category,
            unit='吨', spec='Φ25', standard_price=Decimal('100'),
            safety_stock=Decimal('10'),
        )
        cls.supplier = Supplier.objects.create(
            name='测试供应商', code='SUP_T', contact='张三', phone='13800138000',
        )
        cls.project = Project.objects.create(
            name='测试项目', code='PRJ_T', location='工地A',
        )

        cls.inbound = InboundRecord.objects.create(
            no='IN_T001', material=cls.material,
            quantity=Decimal('50'), unit_price=Decimal('100'),
            supplier=cls.supplier, project=cls.project,
            operator=cls.user, date='2026-03-22',
            location='工地A', spec='Φ25',
        )

        cls.plan = PurchasePlan.objects.create(
            no='PP_T001', project=cls.project, material=cls.material,
            quantity=Decimal('100'), unit_price=Decimal('50'),
            supplier=cls.supplier,
            status=PurchasePlan.STATUS_PURCHASING, operator=cls.user,
        )

        cls.delivery = Delivery.objects.create(
            no='DL_T001', purchase_plan=cls.plan,
            actual_quantity=Decimal('100'), actual_unit_price=Decimal('50'),
            shipping_method='special', plate_number='京A12345',
            supplier=cls.supplier,
            status=Delivery.STATUS_SHIPPED,
        )

    def tearDown(self):
        # 清理导出的临时文件
        export_dir = os.path.join(TEMP_MEDIA, 'exports')
        if os.path.exists(export_dir):
            for f in os.listdir(export_dir):
                os.remove(os.path.join(export_dir, f))


@override_settings(MEDIA_ROOT=TEMP_MEDIA)
class ExportInventoryExcelTest(ExportTaskTestBase):
    """export_inventory_excel 任务测试"""

    def test_export_success(self):
        result = export_inventory_excel(self.user.id)
        self.assertTrue(result['success'])
        self.assertIn('filename', result)
        self.assertTrue(result['filename'].startswith('inventory_summary_'))
        self.assertTrue(os.path.exists(result['file_path']))

    def test_export_file_contains_data(self):
        from openpyxl import load_workbook

        result = export_inventory_excel(self.user.id)
        wb = load_workbook(result['file_path'])
        ws = wb.active
        self.assertEqual(ws.title, '入库汇总')
        # 表头行 + 至少 1 行数据
        self.assertGreaterEqual(ws.max_row, 2)
        # 验证第一行数据包含测试材料信息
        self.assertEqual(ws.cell(row=2, column=1).value, 'MAT_T')
        self.assertEqual(ws.cell(row=2, column=2).value, '测试材料')

    def test_export_with_invalid_user_returns_error(self):
        result = export_inventory_excel(99999)
        self.assertFalse(result['success'])
        self.assertIn('error', result)


@override_settings(MEDIA_ROOT=TEMP_MEDIA)
class ExportInboundExcelTest(ExportTaskTestBase):
    """export_inbound_excel 任务测试"""

    def test_export_success(self):
        result = export_inbound_excel(self.user.id)
        self.assertTrue(result['success'])
        self.assertTrue(os.path.exists(result['file_path']))

    def test_export_with_filters(self):
        from openpyxl import load_workbook

        result = export_inbound_excel(
            self.user.id,
            date_from='2026-03-22',
            date_to='2026-03-22',
            project_id=self.project.pk,
            material_id=self.material.pk,
            supplier_id=self.supplier.pk,
        )
        self.assertTrue(result['success'])
        wb = load_workbook(result['file_path'])
        ws = wb.active
        self.assertEqual(ws.title, '入库记录')
        # 应该有表头行 + 1 行数据
        self.assertEqual(ws.max_row, 2)
        self.assertEqual(ws.cell(row=2, column=1).value, 'IN_T001')

    def test_export_with_date_range_no_match(self):
        from openpyxl import load_workbook

        result = export_inbound_excel(
            self.user.id,
            date_from='2025-01-01',
            date_to='2025-01-02',
        )
        self.assertTrue(result['success'])
        wb = load_workbook(result['file_path'])
        ws = wb.active
        # 只有表头行，无数据
        self.assertEqual(ws.max_row, 1)


@override_settings(MEDIA_ROOT=TEMP_MEDIA)
class ExportPurchasePlansTest(ExportTaskTestBase):
    """export_purchase_plans 任务测试"""

    def test_export_success(self):
        result = export_purchase_plans(self.user.id)
        self.assertTrue(result['success'])
        self.assertTrue(os.path.exists(result['file_path']))

    def test_export_with_status_filter(self):
        from openpyxl import load_workbook

        result = export_purchase_plans(
            self.user.id,
            status=PurchasePlan.STATUS_PURCHASING,
        )
        self.assertTrue(result['success'])
        wb = load_workbook(result['file_path'])
        ws = wb.active
        self.assertEqual(ws.title, '采购计划列表')
        self.assertGreaterEqual(ws.max_row, 2)
        self.assertEqual(ws.cell(row=2, column=1).value, 'PP_T001')

    def test_export_with_search_query(self):
        from openpyxl import load_workbook

        result = export_purchase_plans(
            self.user.id,
            search_query='测试材料',
        )
        self.assertTrue(result['success'])
        wb = load_workbook(result['file_path'])
        ws = wb.active
        self.assertGreaterEqual(ws.max_row, 2)

    def test_export_with_nonexistent_status(self):
        from openpyxl import load_workbook

        result = export_purchase_plans(
            self.user.id,
            status='nonexistent_status',
        )
        self.assertTrue(result['success'])
        wb = load_workbook(result['file_path'])
        ws = wb.active
        # 没有匹配数据，只有表头
        self.assertEqual(ws.max_row, 1)


@override_settings(MEDIA_ROOT=TEMP_MEDIA)
class ExportDeliveriesTest(ExportTaskTestBase):
    """export_deliveries 任务测试"""

    def test_export_success(self):
        result = export_deliveries(self.user.id)
        self.assertTrue(result['success'])
        self.assertTrue(os.path.exists(result['file_path']))

    def test_export_file_contains_delivery_data(self):
        from openpyxl import load_workbook

        result = export_deliveries(self.user.id)
        wb = load_workbook(result['file_path'])
        ws = wb.active
        self.assertEqual(ws.title, '发货单列表')
        self.assertGreaterEqual(ws.max_row, 2)
        # 验证发货单号
        self.assertEqual(ws.cell(row=2, column=1).value, 'DL_T001')
        # 验证采购计划号
        self.assertEqual(ws.cell(row=2, column=2).value, 'PP_T001')
        # 验证数量
        self.assertEqual(ws.cell(row=2, column=7).value, 100.0)
        # 验证单价
        self.assertEqual(ws.cell(row=2, column=8).value, 50.0)
        # 验证总金额
        self.assertEqual(ws.cell(row=2, column=9).value, 5000.0)

    def test_export_with_supplier_filter(self):
        from openpyxl import load_workbook

        result = export_deliveries(
            self.user.id,
            supplier_id=self.supplier.pk,
        )
        self.assertTrue(result['success'])
        wb = load_workbook(result['file_path'])
        ws = wb.active
        self.assertGreaterEqual(ws.max_row, 2)

    def test_export_with_nonexistent_supplier(self):
        from openpyxl import load_workbook

        result = export_deliveries(
            self.user.id,
            supplier_id=99999,
        )
        self.assertTrue(result['success'])
        wb = load_workbook(result['file_path'])
        ws = wb.active
        # 没有匹配数据，只有表头
        self.assertEqual(ws.max_row, 1)

    def test_export_logistics_shipping_method(self):
        """测试物流配送方式的导出"""
        from openpyxl import load_workbook

        Delivery.objects.create(
            no='DL_T002', purchase_plan=self.plan,
            actual_quantity=Decimal('50'), actual_unit_price=Decimal('60'),
            shipping_method='logistics', tracking_no='SF123456',
            supplier=self.supplier,
            status=Delivery.STATUS_PENDING,
        )
        result = export_deliveries(self.user.id)
        self.assertTrue(result['success'])
        wb = load_workbook(result['file_path'])
        ws = wb.active
        # 应至少有 3 行（表头 + 2 条发货单）
        self.assertGreaterEqual(ws.max_row, 3)
