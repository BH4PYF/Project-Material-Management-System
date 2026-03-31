import pytest
from decimal import Decimal
from django.test import TestCase
from django.contrib.auth.models import User
from django.urls import reverse

from ..models import (
    Profile, Category, Material, Supplier,
    Project, InboundRecord, PurchasePlan, Delivery,
)


@pytest.mark.slow
class SlowTestExample(TestCase):
    def test_slow_operation(self):
        import time
        time.sleep(0.1)
        assert True


class DeliveryFlowTestBase(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.admin = User.objects.create_user(username='admin_delivery', password='pass12345')
        Profile.objects.create(user=cls.admin, role='admin')

        cls.material_dept = User.objects.create_user(username='mat_dept', password='pass12345')
        Profile.objects.create(user=cls.material_dept, role='management')

        cls.category = Category.objects.create(name="钢材", code="STEEL_D")
        cls.supplier = Supplier.objects.create(
            name="测试供应商", code="SUP_D01", contact="张三",
            phone="13800138000", main_type=cls.category,
        )
        cls.supplier_user = User.objects.create_user(username='supplier_d', password='pass12345')
        Profile.objects.create(user=cls.supplier_user, role='supplier', supplier_info=cls.supplier)

        cls.clerk = User.objects.create_user(username='clerk_d', password='pass12345')
        Profile.objects.create(user=cls.clerk, role='clerk')

        cls.material = Material.objects.create(
            name="螺纹钢D", code="HRB_D", category=cls.category, unit="吨", spec="Φ25",
        )
        cls.project = Project.objects.create(name="测试项目D", code="P_D01")

        cls.plan = PurchasePlan.objects.create(
            no="PP_D001", project=cls.project, material=cls.material,
            quantity=Decimal('100'), unit_price=Decimal('3800'),
            status=PurchasePlan.STATUS_PURCHASING, operator=cls.admin,
        )

    def setUp(self):
        self.plan.refresh_from_db()


class DeliveryFlowTest(DeliveryFlowTestBase):
    def test_full_delivery_flow(self):
        self.client.login(username='supplier_d', password='pass12345')
        response = self.client.post(reverse('delivery_create'), {
            'purchase_plan_id': self.plan.pk,
            'actual_quantity': '50',
            'actual_unit_price': '3900',
            'shipping_method': 'special',
            'plate_number': '京A12345',
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Delivery.objects.count(), 1)
        delivery = Delivery.objects.first()
        self.assertEqual(delivery.status, 'pending')
        self.assertEqual(delivery.actual_total_amount, Decimal('195000'))

        response = self.client.post(reverse('delivery_confirm_ship', args=[delivery.pk]))
        self.assertEqual(response.status_code, 302)
        delivery.refresh_from_db()
        self.assertEqual(delivery.status, Delivery.STATUS_SHIPPED)
        self.assertIsNotNone(delivery.ship_time)

        self.client.login(username='mat_dept', password='pass12345')
        response = self.client.post(reverse('quick_receive_confirm'), {
            'delivery_id': delivery.pk,
            'receive_date': '2026-03-24',
            'location': '工地A仓库',
        })
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['success'])
        self.assertIn('inbound_no', data)

        self.assertEqual(InboundRecord.objects.count(), 1)
        inbound = InboundRecord.objects.first()
        self.assertEqual(inbound.quantity, Decimal('50'))
        self.assertEqual(inbound.unit_price, Decimal('3900'))

        delivery.refresh_from_db()
        self.assertEqual(delivery.status, 'received')
        self.plan.refresh_from_db()
        self.assertEqual(self.plan.status, 'received')

    def test_create_delivery_for_unapproved_plan_rejected(self):
        self.plan.status = 'pending'
        self.plan.save()
        self.client.login(username='supplier_d', password='pass12345')
        response = self.client.post(reverse('delivery_create'), {
            'purchase_plan_id': self.plan.pk,
            'actual_quantity': '50',
            'actual_unit_price': '3900',
            'shipping_method': 'special',
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Delivery.objects.count(), 0)

    def test_quick_receive_duplicate_rejected(self):
        delivery = Delivery.objects.create(
            no='DL_DUP', purchase_plan=self.plan,
            actual_quantity=Decimal('50'), actual_unit_price=Decimal('3900'),
            shipping_method='special', supplier=self.supplier,
            status='received',
        )
        self.client.login(username='mat_dept', password='pass12345')
        response = self.client.post(reverse('quick_receive_confirm'), {'delivery_id': delivery.pk})
        self.assertEqual(response.status_code, 400)
        self.assertIn('已收货', response.json()['error'])

    def test_get_delivery_by_no(self):
        Delivery.objects.create(
            no='DL_QUERY', purchase_plan=self.plan,
            actual_quantity=Decimal('50'), actual_unit_price=Decimal('3900'),
            shipping_method='special', supplier=self.supplier,
            status=Delivery.STATUS_SHIPPED,
        )
        self.client.login(username='mat_dept', password='pass12345')
        response = self.client.get(reverse('get_delivery_by_no'), {'no': 'DL_QUERY'})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])

    def test_get_delivery_by_no_not_found(self):
        self.client.login(username='mat_dept', password='pass12345')
        response = self.client.get(reverse('get_delivery_by_no'), {'no': 'NONEXIST'})
        self.assertEqual(response.status_code, 404)


class ReportChartAPITest(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.admin = User.objects.create_user(username='report_admin', password='pass12345')
        Profile.objects.create(user=cls.admin, role='admin')

        cls.category = Category.objects.create(name="钢材", code="STEEL_R")
        cls.material = Material.objects.create(
            name="报表测试材料", code="MT_R01", category=cls.category, unit="吨",
        )
        cls.supplier = Supplier.objects.create(
            name="报表测试供应商", code="SUP_R01", contact="王五", phone="13700137000",
        )
        cls.project = Project.objects.create(name="报表测试项目", code="P_R01")

        for i in range(3):
            InboundRecord.objects.create(
                no=f"IN_R{i:03d}", material=cls.material,
                quantity=Decimal('10'), unit_price=Decimal('100'),
                supplier=cls.supplier, project=cls.project,
                operator=cls.admin, date='2026-03-22',
                spec='Φ25',
            )

    def test_chart_data_stock(self):
        self.client.login(username='report_admin', password='pass12345')
        response = self.client.get(reverse('chart_data_api'), {
            'type': 'stock',
            'date_from': '2026-01-01',
            'date_to': '2026-12-31',
        })
        self.assertEqual(response.status_code, 200)
        self.assertIn('data', response.json())

    def test_chart_data_category(self):
        self.client.login(username='report_admin', password='pass12345')
        response = self.client.get(reverse('chart_data_api'), {
            'type': 'category',
            'date_from': '2026-01-01',
            'date_to': '2026-12-31',
        })
        self.assertEqual(response.status_code, 200)
        self.assertIn('data', response.json())

    def test_chart_data_monthly(self):
        self.client.login(username='report_admin', password='pass12345')
        response = self.client.get(reverse('chart_data_api'), {
            'type': 'inbound_monthly',
            'year': '2026',
            'date_from': '2026-01-01',
            'date_to': '2026-12-31',
        })
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertIn('data', data)
        self.assertEqual(len(data['data']), 12)

    def test_get_years_list(self):
        self.client.login(username='report_admin', password='pass12345')
        response = self.client.get(reverse('get_years_list'))
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertIn('years', data)
        self.assertIn(2026, data['years'])

    def test_report_supplier_cost(self):
        self.client.login(username='report_admin', password='pass12345')
        response = self.client.get(reverse('report_supplier_cost'), {
            'supplier_id': self.supplier.pk,
            'date_from': '2026-01-01',
            'date_to': '2026-12-31',
        })
        self.assertEqual(response.status_code, 200)

    def test_report_project_cost(self):
        self.client.login(username='report_admin', password='pass12345')
        response = self.client.get(reverse('report_project_cost'), {
            'project_id': self.project.pk,
            'date_from': '2026-01-01',
            'date_to': '2026-12-31',
        })
        self.assertEqual(response.status_code, 200)

    def test_chart_api_rejects_post(self):
        self.client.login(username='report_admin', password='pass12345')
        response = self.client.post(reverse('chart_data_api'))
        self.assertEqual(response.status_code, 405)


class ExcelExportDataTest(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.admin = User.objects.create_user(username='export_admin', password='pass12345')
        Profile.objects.create(user=cls.admin, role='admin')

        cls.category = Category.objects.create(name="钢材", code="STEEL_E")
        cls.material = Material.objects.create(
            name="导出测试材料", code="MT_E01", category=cls.category,
            unit="吨", standard_price=Decimal('3800'),
        )
        cls.supplier = Supplier.objects.create(
            name="导出测试供应商", code="SUP_E01", contact="赵六", phone="13600136000",
        )
        cls.project = Project.objects.create(name="导出测试项目", code="P_E01")

        InboundRecord.objects.create(
            no="IN_E001", material=cls.material,
            quantity=Decimal('50'), unit_price=Decimal('3800'),
            supplier=cls.supplier, project=cls.project,
            operator=cls.admin, date='2026-03-22',
            spec='Φ25',
        )

    def test_export_inbound_has_data(self):
        from io import BytesIO
        from openpyxl import load_workbook
        self.client.login(username='export_admin', password='pass12345')
        response = self.client.get(reverse('export_excel'), {'type': 'inbound'})
        self.assertEqual(response.status_code, 200)
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        self.assertGreaterEqual(ws.max_row, 2)

    def test_supplier_cost_export_excel(self):
        from io import BytesIO
        from openpyxl import load_workbook
        self.client.login(username='export_admin', password='pass12345')
        response = self.client.get(reverse('report_supplier_cost'), {
            'supplier_id': self.supplier.pk,
            'date_from': '2026-01-01',
            'date_to': '2026-12-31',
            'export': 'excel',
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        self.assertGreaterEqual(ws.max_row, 2)
        amount_cell = ws.cell(row=2, column=7).value
        self.assertEqual(float(amount_cell), 190000.0)
